import io
import os
import re
from datetime import date, datetime
from urllib.parse import quote
from uuid import uuid4

from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.bond import Bond
from app.models.bond_import_log import BondImportLog
from app.models.user import User
from app.schemas.bond import BondImportResult

# Bond type definitions
BOND_TYPES = {
    "A": "일반무담보",
    "B1": "CCRS",
    "B2": "IRL",
    "C": "담보",
}

# Template file mapping
TEMPLATE_FILES = {
    "A": "A_일반무담보_Import_Template.xlsx",
    "B1": "B1_CCRS_Import_Template.xlsx",
    "B2": "B2_IRL_Import_Template.xlsx",
    "C": "C_담보_Import_Template.xlsx",
}

# Common columns across all 4 types → Bond model fields
COMMON_COLUMN_MAP = {
    "자산확정일": "_cutoff_date",
    "금융회사명": "creditor",
    "고객번호": "debtor_id_masked",
    "채권번호": "bond_no",
    "차주 구분": "debtor_type",
    "양도횟수": "transfer_count",
    "상각 여부": "_writeoff",
}

# OPB column name (same across all types)
OPB_COLUMN = "미상환채권잔액(OPB)"

# Type-specific columns that map to Bond model fields
TYPE_SPECIFIC_MAP = {
    "A": {
        "대출상품명": "product_type",
        "최초대출금액": "original_amount",
        "최초 연체일": "overdue_start_date",
    },
    "B1": {
        "연체시작일": "overdue_start_date",
    },
    "B2": {
        "연체시작일": "overdue_start_date",
    },
    "C": {
        "자산유형": "collateral_type",
        "대출과목": "product_type",
        "최초대출원금": "original_amount",
        "연체시작일": "overdue_start_date",
        "채권원금잔액": "interest_balance",
    },
}

# Fields that should be parsed as integers
INT_FIELDS = {"original_amount", "opb", "interest_balance", "total_balance", "transfer_count"}

# Fields that should be parsed as dates
DATE_FIELDS = {"overdue_start_date", "_cutoff_date"}

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "template")


def get_template(bond_type: str) -> StreamingResponse:
    """Serve a pre-made Excel template file for the given bond type."""
    if bond_type not in TEMPLATE_FILES:
        raise HTTPException(400, f"지원하지 않는 채권유형입니다: {bond_type}")

    filename = TEMPLATE_FILES[bond_type]
    filepath = os.path.join(TEMPLATE_DIR, filename)

    if not os.path.exists(filepath):
        raise HTTPException(404, f"템플릿 파일을 찾을 수 없습니다: {filename}")

    with open(filepath, "rb") as f:
        content = f.read()

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# Keep backward compat
def generate_template() -> StreamingResponse:
    """Default: serve type A template."""
    return get_template("A")


def _parse_value(val, model_field: str):
    """Parse a cell value for the appropriate field type."""
    if val is None:
        return None

    if model_field in DATE_FIELDS:
        if isinstance(val, (date, datetime)):
            return val if isinstance(val, date) and not isinstance(val, datetime) else val.date()
        # Excel serial number (int or float) → date
        if isinstance(val, (int, float)) and 1 < val < 100000:
            from datetime import timedelta
            excel_epoch = date(1899, 12, 30)
            return excel_epoch + timedelta(days=int(val))
        val_str = str(val).strip()
        if not val_str:
            return None
        return date.fromisoformat(val_str)

    if model_field in INT_FIELDS:
        if isinstance(val, (int, float)):
            return int(val)
        val_str = str(val).strip()
        if not val_str:
            return None
        # Extract numeric part (e.g., "최초 매각 = 0" → "0")
        cleaned = val_str.replace(",", "")
        try:
            return int(float(cleaned))
        except (ValueError, OverflowError):
            # Try extracting the last number from the string
            nums = re.findall(r'-?\d+\.?\d*', cleaned)
            return int(float(nums[-1])) if nums else None

    val_str = str(val).strip()
    return val_str if val_str else None


async def import_excel(
    file: UploadFile,
    pool_id: int,
    user: User,
    db: AsyncSession,
    bond_type: str = "A",
) -> BondImportResult:
    if bond_type not in BOND_TYPES:
        raise HTTPException(400, f"지원하지 않는 채권유형입니다: {bond_type}")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return BondImportResult(
            file_name=file.filename or "unknown.xlsx",
            row_count=0, success_count=0, error_count=0, errors=None,
        )

    # First row = headers
    headers = [re.sub(r'\s+', ' ', str(h)).strip() if h else "" for h in rows[0]]

    # Build combined column map for this bond_type
    combined_map = dict(COMMON_COLUMN_MAP)
    combined_map[OPB_COLUMN] = "opb"
    type_map = TYPE_SPECIFIC_MAP.get(bond_type, {})
    combined_map.update(type_map)

    # Build header index → (model_field, is_extra) mapping
    header_map: dict[int, tuple[str, bool]] = {}
    model_fields = set(combined_map.values())
    for idx, h in enumerate(headers):
        if h in combined_map:
            field = combined_map[h]
            header_map[idx] = (field, False)
        elif h:
            # Store all other columns in extra_data
            header_map[idx] = (h, True)

    batch_id = str(uuid4())[:8]
    bonds: list[Bond] = []
    errors: list[dict] = []
    row_count = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip description/guide row (row 2) — check if first value looks like sample data
        if row_idx == 2:
            # If first cell is a date or has sample-like content, it's likely data, not a guide row
            first_val = row[0] if row[0] else None
            if first_val is None or str(first_val).strip() == "":
                continue
            # Check if it looks like a guide row (text description, not a real date)
            first_str = str(first_val).strip()
            if first_str.startswith("예:") or first_str.startswith("("):
                continue
            # Otherwise it's a data row — process it below

        # Skip fully empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_count += 1
        try:
            bond_data: dict = {
                "pool_id": pool_id,
                "bond_type": bond_type,
                "import_batch": batch_id,
                "created_by": user.id,
            }
            extra_data: dict = {}

            for col_idx, (field, is_extra) in header_map.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None or str(val).strip() == "":
                    continue

                if is_extra:
                    # Store in extra_data JSON
                    if isinstance(val, (date, datetime)):
                        extra_data[field] = val.isoformat() if isinstance(val, date) and not isinstance(val, datetime) else val.date().isoformat()
                    else:
                        val_str = str(val).strip()
                        # Try to convert numeric strings
                        try:
                            num = float(val_str.replace(",", ""))
                            extra_data[field] = int(num) if num == int(num) else num
                        except (ValueError, OverflowError):
                            extra_data[field] = val_str
                else:
                    # Skip internal fields (prefixed with _)
                    if field.startswith("_"):
                        continue
                    bond_data[field] = _parse_value(val, field)

            if extra_data:
                bond_data["extra_data"] = extra_data

            bonds.append(Bond(**bond_data))
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    wb.close()

    # Bulk insert valid bonds
    db.add_all(bonds)

    # Create import log
    log = BondImportLog(
        pool_id=pool_id,
        file_name=file.filename or "unknown.xlsx",
        row_count=row_count,
        success_count=len(bonds),
        error_count=len(errors),
        errors=errors if errors else None,
        imported_by=user.id,
    )
    db.add(log)
    await db.commit()

    return BondImportResult(
        file_name=file.filename or "unknown.xlsx",
        row_count=row_count,
        success_count=len(bonds),
        error_count=len(errors),
        errors=errors if errors else None,
    )
